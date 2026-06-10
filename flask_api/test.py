from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import random

random.seed(42)

# Real Bangalore localities with realistic street-level addresses
bangalore_addresses = [
    # Koramangala
    ("14, 5th Cross, Koramangala 4th Block, Bangalore - 560034", "Koramangala"),
    ("27, 80 Feet Rd, Koramangala 6th Block, Bangalore - 560095", "Koramangala"),
    ("3, 1st Main Rd, Koramangala 3rd Block, Bangalore - 560034", "Koramangala"),
    ("88, Sarjapur Rd, Koramangala 1st Block, Bangalore - 560034", "Koramangala"),
    ("45, 7th Cross, Koramangala 5th Block, Bangalore - 560095", "Koramangala"),
    ("12, 2nd Main Rd, Koramangala 8th Block, Bangalore - 560095", "Koramangala"),
    # Indiranagar
    ("34, 100 Feet Rd, Indiranagar, Bangalore - 560038", "Indiranagar"),
    ("7, 12th Main Rd, Indiranagar, Bangalore - 560038", "Indiranagar"),
    ("56, CMH Rd, Indiranagar, Bangalore - 560038", "Indiranagar"),
    ("21, 5th Cross, Indiranagar 1st Stage, Bangalore - 560038", "Indiranagar"),
    ("9, Defence Colony, Indiranagar, Bangalore - 560038", "Indiranagar"),
    ("18, HAL 2nd Stage, Indiranagar, Bangalore - 560038", "Indiranagar"),
    # Whitefield
    ("42, Whitefield Main Rd, Whitefield, Bangalore - 560066", "Whitefield"),
    ("15, ITPL Main Rd, Whitefield, Bangalore - 560066", "Whitefield"),
    ("73, Varthur Rd, Whitefield, Bangalore - 560066", "Whitefield"),
    ("8, Marathahalli - Sarjapur Outer Ring Rd, Whitefield, Bangalore - 560066", "Whitefield"),
    ("29, Hope Farm Junction, Whitefield, Bangalore - 560066", "Whitefield"),
    ("61, Kadugodi Main Rd, Whitefield, Bangalore - 560067", "Whitefield"),
    # HSR Layout
    ("19, 27th Main Rd, HSR Layout Sector 1, Bangalore - 560102", "HSR Layout"),
    ("5, 24th Main Rd, HSR Layout Sector 2, Bangalore - 560102", "HSR Layout"),
    ("38, BDA Complex, HSR Layout, Bangalore - 560102", "HSR Layout"),
    ("11, 22nd Cross, HSR Layout Sector 7, Bangalore - 560102", "HSR Layout"),
    ("66, Outer Ring Rd, HSR Layout, Bangalore - 560102", "HSR Layout"),
    # Marathahalli
    ("23, Marathahalli Bridge, Marathahalli, Bangalore - 560037", "Marathahalli"),
    ("47, Doddanekundi, Marathahalli, Bangalore - 560037", "Marathahalli"),
    ("6, Kundalahalli Gate, Marathahalli, Bangalore - 560037", "Marathahalli"),
    ("82, Spice Garden Layout, Marathahalli, Bangalore - 560037", "Marathahalli"),
    # Electronic City
    ("33, Electronic City Phase 1, Bangalore - 560100", "Electronic City"),
    ("17, Electronic City Phase 2, Bangalore - 560100", "Electronic City"),
    ("54, Neeladri Rd, Electronic City, Bangalore - 560100", "Electronic City"),
    ("2, Hebbagodi, Electronic City, Bangalore - 560099", "Electronic City"),
    ("70, Konappana Agrahara, Electronic City, Bangalore - 560100", "Electronic City"),
    # Jayanagar
    ("10, 11th Cross, Jayanagar 3rd Block, Bangalore - 560011", "Jayanagar"),
    ("44, 9th Main Rd, Jayanagar 4th Block, Bangalore - 560041", "Jayanagar"),
    ("26, 30th Cross, Jayanagar 7th Block, Bangalore - 560082", "Jayanagar"),
    ("58, South End Rd, Jayanagar, Bangalore - 560011", "Jayanagar"),
    # Rajajinagar
    ("13, 8th Cross, Rajajinagar 1st Block, Bangalore - 560010", "Rajajinagar"),
    ("37, 3rd Main Rd, Rajajinagar 2nd Block, Bangalore - 560010", "Rajajinagar"),
    ("75, Chord Rd, Rajajinagar, Bangalore - 560010", "Rajajinagar"),
    ("4, Nehru Circle, Rajajinagar, Bangalore - 560010", "Rajajinagar"),
    # Malleswaram
    ("20, 15th Cross, Malleswaram, Bangalore - 560003", "Malleswaram"),
    ("63, Sampige Rd, Malleswaram, Bangalore - 560003", "Malleswaram"),
    ("39, 11th Main Rd, Malleswaram, Bangalore - 560055", "Malleswaram"),
    # Banashankari
    ("16, 4th Stage, Banashankari, Bangalore - 560085", "Banashankari"),
    ("52, Kanakapura Rd, Banashankari 2nd Stage, Bangalore - 560070", "Banashankari"),
    ("31, BSK 3rd Stage, Banashankari, Bangalore - 560085", "Banashankari"),
    # Bellandur
    ("48, Bellandur Gate, Outer Ring Rd, Bellandur, Bangalore - 560103", "Bellandur"),
    ("9, Kadubeesanahalli, Bellandur, Bangalore - 560103", "Bellandur"),
    ("72, Sarjapur Rd, Bellandur, Bangalore - 560103", "Bellandur"),
    # Hebbal
    ("25, Hebbal Kempapura, Hebbal, Bangalore - 560024", "Hebbal"),
    ("41, Outer Ring Rd, Hebbal, Bangalore - 560024", "Hebbal"),
    ("14, MS Ramaiah Rd, Hebbal, Bangalore - 560024", "Hebbal"),
    # JP Nagar
    ("36, 24th Main Rd, JP Nagar Phase 1, Bangalore - 560078", "JP Nagar"),
    ("60, Uttarahalli Rd, JP Nagar Phase 6, Bangalore - 560078", "JP Nagar"),
    ("8, JP Nagar 7th Phase, Bangalore - 560078", "JP Nagar"),
    # Yelahanka
    ("22, Yelahanka New Town, Bangalore - 560064", "Yelahanka"),
    ("55, Kogilu Main Rd, Yelahanka, Bangalore - 560064", "Yelahanka"),
    ("11, Doddaballapur Rd, Yelahanka, Bangalore - 560064", "Yelahanka"),
    # Yeshwanthpur
    ("43, Tumkur Rd, Yeshwanthpur, Bangalore - 560022", "Yeshwanthpur"),
    ("77, Goraguntepalya, Yeshwanthpur, Bangalore - 560022", "Yeshwanthpur"),
    # Bannerghatta Road
    ("18, Bannerghatta Rd, JP Nagar 9th Phase, Bangalore - 560076", "Bannerghatta Road"),
    ("67, Gottigere, Bannerghatta Rd, Bangalore - 560083", "Bannerghatta Road"),
    ("32, Arekere, Bannerghatta Rd, Bangalore - 560076", "Bannerghatta Road"),
    # Sarjapur Road
    ("50, Sarjapur Rd, Carmelaram, Bangalore - 560035", "Sarjapur Road"),
    ("28, Ambalipura, Sarjapur Rd, Bangalore - 560102", "Sarjapur Road"),
    ("85, Harlur Rd, Sarjapur, Bangalore - 560102", "Sarjapur Road"),
    # Vijayanagar
    ("16, 4th Main Rd, Vijayanagar, Bangalore - 560040", "Vijayanagar"),
    ("49, Chord Rd, Vijayanagar 1st Stage, Bangalore - 560040", "Vijayanagar"),
]

first_names = [
    "Rajesh","Priya","Amit","Sneha","Karthik","Deepa","Suresh","Anita","Vijay","Lakshmi",
    "Rahul","Pooja","Arun","Divya","Sanjay","Meera","Ravi","Kavya","Manoj","Swati",
    "Naveen","Shweta","Girish","Nandita","Prasad","Asha","Vinay","Rekha","Ajay","Usha",
    "Harish","Sindhu","Nikhil","Bhavna","Ramesh","Leela","Sunil","Gayatri","Dinesh","Saritha",
    "Kishore","Padma","Venkat","Jyothi","Murali","Hema","Ashok","Vani","Pradeep","Geetha",
    "Rohit","Suma","Sachin","Nirmala","Arjun","Shanthi","Vikram","Radha","Ganesh","Lalitha",
    "Santosh","Manjula","Praveen","Savitha","Shivanand","Nalini","Balaji","Kamala","Mahesh","Sudha"
]
last_names = [
    "Kumar","Sharma","Reddy","Naidu","Gowda","Patel","Rao","Nair","Iyer","Hegde",
    "Joshi","Pillai","Menon","Shetty","Kamath","Bhat","Krishnamurthy","Venkatesh","Subramanian","Rajan"
]

shifts = ["09:00", "09:00", "09:00", "10:00", "10:00", "08:00"]  # weighted toward 9am

employees = []
used_combos = set()
for i in range(50):
    while True:
        fn = random.choice(first_names)
        ln = random.choice(last_names)
        if (fn, ln) not in used_combos:
            used_combos.add((fn, ln))
            break
    addr_data = bangalore_addresses[i % len(bangalore_addresses)]
    phone = f"9{random.randint(100000000, 999999999)}"
    shift = random.choice(shifts)
    shift_type = {"09:00": "Morning", "10:00": "Morning", "08:00": "Early Morning"}.get(shift, "Morning")
    employees.append({
        "employee_id": f"EMP{str(i+1).zfill(3)}",
        "name": f"{fn} {ln}",
        "address": addr_data[0],
        "area": addr_data[1],
        "shift_time": shift,
        "shift_type": shift_type,
        "phone": phone,
    })

wb = Workbook()

# ── Sheet 1: Employee Data ─────────────────────────────────────────────
ws = wb.active
ws.title = "Employee Data"

# Styles
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", start_color="1A3C5E")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
data_font = Font(name="Arial", size=10)
alt_fill = PatternFill("solid", start_color="F0F5FB")
white_fill = PatternFill("solid", start_color="FFFFFF")
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="CCCCCC")
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = ["employee_id", "name", "address", "area", "shift_time", "shift_type", "phone"]
col_widths = [14, 22, 60, 18, 12, 16, 14]

# Title row
ws.merge_cells("A1:G1")
ws["A1"] = "CabRoute Optimizer — Sample Employee Dataset (Bangalore)"
ws["A1"].font = Font(name="Arial", bold=True, size=13, color="1A3C5E")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws["A1"].fill = PatternFill("solid", start_color="DCE9F5")
ws.row_dimensions[1].height = 28

# Sub-note
ws.merge_cells("A2:G2")
ws["A2"] = "50 sample employees across 15+ Bangalore localities  |  Use this as the upload template for CabRoute Optimizer"
ws["A2"].font = Font(name="Arial", size=9, italic=True, color="666666")
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 18

# Header row (row 3)
for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col)].width = w
ws.row_dimensions[3].height = 22

# Data rows
for row_idx, emp in enumerate(employees, 4):
    fill = alt_fill if row_idx % 2 == 0 else white_fill
    values = [emp["employee_id"], emp["name"], emp["address"], emp["area"],
              emp["shift_time"], emp["shift_type"], emp["phone"]]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.font = data_font
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = center if col in [1, 5, 6, 7] else left
ws.row_dimensions[row_idx].height = 30  # taller for address wrap

ws.freeze_panes = "A4"

# ── Sheet 2: Format Guide ──────────────────────────────────────────────
ws2 = wb.create_sheet("Format Guide")
ws2.column_dimensions["A"].width = 18
ws2.column_dimensions["B"].width = 20
ws2.column_dimensions["C"].width = 14
ws2.column_dimensions["D"].width = 40

# Title
ws2.merge_cells("A1:D1")
ws2["A1"] = "Column Format Reference"
ws2["A1"].font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
ws2["A1"].fill = PatternFill("solid", start_color="1A3C5E")
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 26

guide_headers = ["Column", "Type", "Required", "Notes / Example"]
for col, h in enumerate(guide_headers, 1):
    cell = ws2.cell(row=2, column=col, value=h)
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", start_color="2E6DA4")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

guide_rows = [
    ("employee_id", "Text / Number", "YES", "Unique ID. e.g. EMP001, 1001, A-045"),
    ("name",        "Text",          "YES", "Full name. e.g. Rajesh Kumar"),
    ("address",     "Text",          "YES", "Full address with locality & city. Needed for geocoding accuracy."),
    ("area",        "Text",          "YES", "Locality name only. e.g. Koramangala, Whitefield"),
    ("shift_time",  "HH:MM (24hr)",  "YES", "24-hour format only. e.g. 09:00, 18:00, 22:30"),
    ("shift_type",  "Text",          "NO",  "Morning / Evening / Night. Auto-derived if left blank."),
    ("phone",       "Text",          "NO",  "10-digit mobile. Store as text to preserve leading digits."),
]

for row_idx, row_data in enumerate(guide_rows, 3):
    fill = alt_fill if row_idx % 2 == 0 else white_fill
    for col, val in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col, value=val)
        cell.font = Font(name="Arial", size=10)
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws2.row_dimensions[row_idx].height = 30

ws2.cell(row=11, column=1).value = ""
ws2.merge_cells("A12:D12")
ws2["A12"] = "IMPORTANT: Do not rename or reorder column headers. Extra columns are ignored. First row of data must start at row 1 (no blank rows above headers)."
ws2["A12"].font = Font(name="Arial", size=9, italic=True, color="CC0000")
ws2["A12"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws2.row_dimensions[12].height = 36

# ── Sheet 3: Config ────────────────────────────────────────────────────
ws3 = wb.create_sheet("Config")
ws3.column_dimensions["A"].width = 28
ws3.column_dimensions["B"].width = 16
ws3.column_dimensions["C"].width = 35

ws3.merge_cells("A1:C1")
ws3["A1"] = "Default Configuration"
ws3["A1"].font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
ws3["A1"].fill = PatternFill("solid", start_color="1A3C5E")
ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 26

for col, h in enumerate(["Parameter", "Default Value", "Description"], 1):
    cell = ws3.cell(row=2, column=col, value=h)
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", start_color="2E6DA4")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

config_rows = [
    ("cab_capacity",        "6",              "Max employees per cab"),
    ("rate_per_km",         "18",             "Cost in Rs. per km"),
    ("working_days_month",  "26",             "Working days used for monthly cost estimate"),
    ("office_lat",          "12.9352",        "Office latitude (default: Koramangala)"),
    ("office_lng",          "77.6245",        "Office longitude (default: Koramangala)"),
    ("office_address",      "L&T Finance, Koramangala, Bangalore", "Office address label on map"),
    ("geocode_cache",       "TRUE",           "Cache geocoded addresses to avoid repeat API calls"),
    ("max_employees",       "1000",           "Max rows allowed per upload"),
]

for row_idx, row_data in enumerate(config_rows, 3):
    fill = alt_fill if row_idx % 2 == 0 else white_fill
    for col, val in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col, value=val)
        cell.font = Font(name="Arial", size=10)
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws3.row_dimensions[row_idx].height = 22

wb.save("cabroute_sample_dataset.xlsx")
print("Done")