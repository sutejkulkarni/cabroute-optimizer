"""
excel_reader.py
Reads employee data from the uploaded Excel file.
Validates required columns and normalizes data.
"""

import openpyxl

REQUIRED_COLUMNS = {"employee_id", "name", "address", "area", "shift_time"}


def read_employees(filepath: str) -> list[dict]:
    """
    Reads employee data from .xlsx file.
    Returns list of employee dicts.
    Raises ValueError on missing columns or empty file.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Try 'Employee Data' sheet first, fall back to first sheet
    if "Employee Data" in wb.sheetnames:
        ws = wb["Employee Data"]
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))

    # Find header row (first non-empty row)
    header_row_idx = None
    headers = []
    for i, row in enumerate(rows):
        non_empty = [str(c).strip().lower() for c in row if c is not None]
        if REQUIRED_COLUMNS.issubset(set(non_empty)):
            header_row_idx = i
            headers = [str(c).strip().lower() if c else "" for c in row]
            break

    if header_row_idx is None:
        raise ValueError(
            f"Could not find required headers. Expected: {REQUIRED_COLUMNS}. "
            "Check your Excel file matches the template format."
        )

    employees = []
    for row in rows[header_row_idx + 1:]:
        if all(c is None for c in row):
            continue  # skip blank rows

        emp = {headers[i]: (str(row[i]).strip() if row[i] is not None else "") for i in range(len(headers)) if headers[i]}

        # Skip rows with missing required fields
        if not all(emp.get(col) for col in REQUIRED_COLUMNS):
            continue

        # Normalize shift_time to HH:MM
        emp["shift_time"] = _normalize_shift(emp.get("shift_time", "09:00"))
        employees.append(emp)

    if not employees:
        raise ValueError("No valid employee rows found in the file.")

    print(f"[excel_reader] Loaded {len(employees)} employees from '{filepath}'")
    return employees


def _normalize_shift(val: str) -> str:
    """Ensures shift_time is in HH:MM format."""
    val = str(val).strip()
    if ":" in val:
        parts = val.split(":")
        return f"{int(parts[0]):02d}:{parts[1][:2]}"
    try:
        h = int(val)
        return f"{h:02d}:00"
    except Exception:
        return "09:00"
